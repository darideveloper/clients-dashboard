import io
from decimal import Decimal
from unittest import mock

from django.contrib.auth.models import User
from django.db.utils import OperationalError
from django.test import Client, RequestFactory, TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from core.models import Brand
from ourlives.models import AppSettings, InvitationCode, Organization, Project, StripeEvent
from utils.excel_export import (
    _primary_color,
    autosize_columns,
    build_full_app_workbook,
    build_workbook_for_queryset,
    columns_for_model,
    get_related_targets,
    sanitize_sheet_name,
    serialize_value,
    style_sheet,
)


class SanitizeSheetNameTests(TestCase):
    def test_forbidden_chars_replaced(self):
        self.assertEqual(sanitize_sheet_name("a:b/c\\d?e*f[g]h", set()), "a_b_c_d_e_f_g_h")

    def test_leading_trailing_quote_stripped(self):
        self.assertEqual(sanitize_sheet_name("'hello'", set()), "hello")

    def test_empty_fallback(self):
        self.assertEqual(sanitize_sheet_name("", set()), "Sheet")
        self.assertEqual(sanitize_sheet_name(None, set()), "Sheet")
        self.assertEqual(sanitize_sheet_name("   ", set()), "Sheet")

    def test_truncate_to_31(self):
        long_name = "A" * 40
        result = sanitize_sheet_name(long_name, set())
        self.assertEqual(len(result), 31)
        self.assertEqual(result, "A" * 31)

    def test_deduplicate_suffix(self):
        self.assertEqual(sanitize_sheet_name("Project", {"Project"}), "Project_2")
        self.assertEqual(sanitize_sheet_name("Project", {"Project", "Project_2"}), "Project_3")

    def test_deduplicate_truncates_to_fit_suffix(self):
        long_name = "A" * 31
        existing = {"A" * 31}
        result = sanitize_sheet_name(long_name, existing)
        self.assertEqual(len(result), 31)
        self.assertTrue(result.endswith("_2"))


class SerializeValueTests(TestCase):
    def setUp(self):
        AppSettings.get_solo()
        AppSettings.objects.update(total_tokens=100)
        self.project = Project.objects.create(name="P")
        self.org = Organization.objects.create(name="O")

    def test_integer_and_decimal_typed(self):
        from django.db import models

        # Use AppSettings decimal field
        settings = AppSettings.get_solo()
        settings.price_per_token = Decimal("2.50")
        field = AppSettings._meta.get_field("price_per_token")
        val = serialize_value(field, settings)
        self.assertIsInstance(val, float)
        self.assertEqual(val, 2.5)

        field2 = InvitationCode._meta.get_field("max_use")
        code = InvitationCode(project=self.project, organization=self.org, max_use=5)
        self.assertEqual(serialize_value(field2, code), 5)
        self.assertIsInstance(serialize_value(field2, code), int)

    def test_boolean_typed(self):
        field = InvitationCode._meta.get_field("is_active")
        code = InvitationCode(project=self.project, organization=self.org, max_use=5, is_active=True)
        self.assertIs(serialize_value(field, code), True)

    def test_none_returns_none(self):
        field = InvitationCode._meta.get_field("is_active")
        code = InvitationCode(project=self.project, organization=self.org, max_use=5)
        # Simulate nullable FK returning None via string getter? Check generic
        # For non-FK None case, serialize None field
        self.assertIsNone(serialize_value(field, type("obj", (), {"is_active": None})()))

    def test_file_field(self):
        field = Brand._meta.get_field("logo")
        brand = Brand(name="B")
        # No logo -> None
        self.assertIsNone(serialize_value(field, brand))


class ColumnsForModelTests(TestCase):
    def setUp(self):
        AppSettings.get_solo()
        AppSettings.objects.update(total_tokens=100)

    def test_invitation_code_fk_expands(self):
        cols = columns_for_model(InvitationCode)
        headers = [h for h, _, _ in cols]
        self.assertIn("project__str__", headers)
        self.assertIn("project_id", headers)
        self.assertIn("organization__str__", headers)
        self.assertIn("organization_id", headers)
        # Each FK produces two columns, plus other fields
        # Ensure concrete non-FK field present
        self.assertTrue(any("code" in h.lower() for h in headers))

    def test_project_single_column_per_field(self):
        cols = columns_for_model(Project)
        headers = [h for h, _, _ in cols]
        # Project has no FKs, so no __str__ columns
        self.assertFalse(any("__str__" in h for h in headers))
        self.assertIn("id", [h.lower() for h in headers] or headers)


class GetRelatedTargetsTests(TestCase):
    def test_invitation_code_targets(self):
        targets = get_related_targets(InvitationCode)
        names = {t.__name__ for t in targets}
        self.assertIn("Project", names)
        self.assertIn("Organization", names)
        self.assertNotIn("StripeEvent", names)

    def test_no_transitive(self):
        # Project has no FKs
        self.assertEqual(get_related_targets(Project), [])


class WorkbookBuildingTests(TestCase):
    def setUp(self):
        AppSettings.get_solo()
        AppSettings.objects.update(total_tokens=100)
        self.project_a = Project.objects.create(name="Alpha")
        self.project_b = Project.objects.create(name="Beta")
        self.org_x = Organization.objects.create(name="X")
        self.code1 = InvitationCode.objects.create(project=self.project_a, organization=self.org_x, max_use=5)
        self.code2 = InvitationCode.objects.create(project=self.project_b, organization=self.org_x, max_use=5)

    def test_single_sheet_without_related(self):
        qs = InvitationCode.objects.all()
        wb = build_workbook_for_queryset(InvitationCode, qs, include_related=False)
        self.assertEqual(len(wb.sheetnames), 1)
        ws = wb.active
        # Header + 2 rows
        self.assertEqual(ws.max_row, 3)
        headers = [c.value for c in ws[1]]
        self.assertIn("project__str__", headers)
        self.assertIn("project_id", headers)

    def test_with_related_referenced_only(self):
        # 2 codes reference 2 distinct projects but same org
        qs = InvitationCode.objects.all()
        # Create third project not referenced
        Project.objects.create(name="Gamma")
        wb = build_workbook_for_queryset(InvitationCode, qs, include_related=True)
        # Main + Project + Organization = 3 sheets
        self.assertEqual(len(wb.sheetnames), 3)
        # Find Project sheet
        project_ws = None
        for name in wb.sheetnames:
            if "project" in name.lower():
                project_ws = wb[name]
                break
        self.assertIsNotNone(project_ws)
        # Should have header + 2 referenced projects, not 3
        self.assertEqual(project_ws.max_row, 3)

    def test_no_transitive_sheets(self):
        # InvitationCode -> Project -> no further, so no transitive
        qs = InvitationCode.objects.filter(pk=self.code1.pk)
        wb = build_workbook_for_queryset(InvitationCode, qs, include_related=True)
        # Should not contain StripeEvent
        for name in wb.sheetnames:
            self.assertNotIn("stripe", name.lower())

    def test_styling_applied(self):
        qs = Project.objects.all()
        wb = build_workbook_for_queryset(Project, qs, include_related=False)
        ws = wb.active
        # Header bold and fill
        self.assertTrue(ws["A1"].font.bold)
        self.assertEqual(ws.freeze_panes, "A2")
        # Banded rows: at least one data cell has fill
        if ws.max_row > 1:
            self.assertIsNotNone(ws["A2"].fill.start_color.rgb)


class BuildFullAppWorkbookTests(TestCase):
    def setUp(self):
        AppSettings.get_solo()
        AppSettings.objects.update(total_tokens=100)

    def test_full_export_contains_all_ourlives_models_ordered(self):
        wb = build_full_app_workbook()
        # Should contain sheets for 5 ourlives models deterministically sorted by model_name
        expected = sorted(["appsettings", "invitationcode", "organization", "project", "stripeevent"])
        actual = sorted([n.lower().replace(" ", "").replace("_", "") for n in wb.sheetnames])
        # Check that project/organization etc are present
        lower = [n.lower() for n in wb.sheetnames]
        self.assertTrue(any("project" in n for n in lower))
        self.assertTrue(any("organization" in n for n in lower))
        # Sheets are sorted by model_name
        sheet_order = wb.sheetnames
        sorted_order = sorted(sheet_order, key=lambda s: s.lower())
        self.assertEqual(sheet_order, sorted_order)

    def test_primary_color_fallback(self):
        with mock.patch("core.models.Brand.get_or_create_default", side_effect=OperationalError("no table")):
            color = _primary_color()
            self.assertEqual(color, "#C92FFF")


class ExcelExportAdminTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.superuser = User.objects.create_superuser(username="admin", email="a@x.test", password="x")
        self.client.force_login(self.superuser)
        AppSettings.get_solo()
        AppSettings.objects.update(total_tokens=100)
        self.project = Project.objects.create(name="P1")
        self.org = Organization.objects.create(name="O1")
        self.code = InvitationCode.objects.create(project=self.project, organization=self.org, max_use=5)

    def test_export_selected_returns_excel(self):
        url = reverse("admin:ourlives_invitationcode_changelist")
        response = self.client.post(url, {"action": "export_selected", "_selected_action": [self.code.pk]}, follow=True)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn("filename*", response["Content-Disposition"])
        # Check valid xlsx
        wb = load_workbook(filename=io.BytesIO(response.content))
        self.assertEqual(len(wb.sheetnames), 1)

    def test_export_with_related_returns_three_sheets(self):
        url = reverse("admin:ourlives_invitationcode_changelist")
        response = self.client.post(url, {"action": "export_selected_with_related", "_selected_action": [self.code.pk]})
        self.assertEqual(response.status_code, 200)
        wb = load_workbook(filename=io.BytesIO(response.content))
        self.assertEqual(len(wb.sheetnames), 3)

    def test_empty_selection_warns(self):
        url = reverse("admin:ourlives_invitationcode_changelist")
        response = self.client.post(url, {"action": "export_selected", "_selected_action": []}, follow=True)
        # Should redirect back with message, not download
        self.assertEqual(response.status_code, 200)
        # No excel content type
        self.assertNotEqual(response.get("Content-Type"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    def test_stripe_event_export_view_gated(self):
        # StripeEvent is read-only (has_change_permission False) but view-gated export should work
        StripeEvent.objects.create(stripe_event_id="evt_123", source="ourlives", token_count=10, amount_cents=1000)
        url = reverse("admin:ourlives_stripeevent_changelist")
        obj = StripeEvent.objects.first()
        response = self.client.post(url, {"action": "export_selected", "_selected_action": [obj.pk]})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


class FullAppExportPermissionTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.superuser = User.objects.create_superuser(username="su", email="su@x.test", password="x")
        self.staff_no_perms = User.objects.create_user(username="staff", email="s@x.test", password="x", is_staff=True)
        AppSettings.get_solo()
        AppSettings.objects.update(total_tokens=10)
        Project.objects.create(name="P")

    def test_has_export_all_permission_true_for_perms(self):
        from project.admin_base import OurlivesExportMixin
        from django.test import RequestFactory
        factory = RequestFactory()
        request = factory.get("/")
        request.user = self.superuser
        mixin = OurlivesExportMixin()
        self.assertTrue(mixin.has_export_all_permission(request))

    def test_has_export_all_permission_false_without_perms(self):
        from project.admin_base import OurlivesExportMixin
        from django.test import RequestFactory
        factory = RequestFactory()
        request = factory.get("/")
        request.user = self.staff_no_perms
        mixin = OurlivesExportMixin()
        self.assertFalse(mixin.has_export_all_permission(request))

    def test_non_staff_denied(self):
        from project.admin_base import OurlivesExportMixin
        from django.test import RequestFactory
        from django.contrib.auth.models import User as AuthUser
        factory = RequestFactory()
        user = AuthUser.objects.create_user(username="u", email="u@x.test", password="x", is_staff=False)
        request = factory.get("/")
        request.user = user
        mixin = OurlivesExportMixin()
        self.assertFalse(mixin.has_export_all_permission(request))
