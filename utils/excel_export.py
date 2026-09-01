"""
Excel export utility for admin.

Provides workbook building with proper formatting, FK handling,
and related-data sheets. See specs/excel-export/spec.md.
"""
import json
from datetime import date, datetime, time

from django.apps import apps
from django.db import models
from django.db.utils import OperationalError, ProgrammingError
from django.utils.text import slugify


FALLBACK_COLOR = "#C92FFF"


def _primary_color():
    """Return brand primary color or fallback; guards missing table."""
    try:
        from core.models import Brand

        brand = Brand.get_or_create_default()
        color = getattr(brand, "primary_color", None)
        if color:
            return color
    except (OperationalError, ProgrammingError):
        pass
    except Exception:
        pass
    return FALLBACK_COLOR


def sanitize_sheet_name(name, existing=None):
    """
    Sanitize Excel sheet name.

    - Replaces : \\ / ? * [ ] with _
    - Strips leading/trailing '
    - Falls back to model_name if empty
    - Truncates to 31 chars
    - Deduplicates with _2, _3 suffix
    """
    existing_set = set(existing) if existing else set()
    raw = str(name) if name is not None else ""
    raw = raw.strip()
    if not raw:
        raw = "Sheet"

    # Replace forbidden chars
    forbidden = set(r':\/?*[]')
    sanitized = "".join("_" if c in forbidden else c for c in raw)
    # Strip leading/trailing '
    sanitized = sanitized.strip("'")
    if not sanitized:
        sanitized = "Sheet"

    # Truncate to 31
    sanitized = sanitized[:31]

    if sanitized not in existing_set:
        return sanitized

    # Deduplicate
    for i in range(2, 1000):
        suffix = f"_{i}"
        base_len = 31 - len(suffix)
        candidate = sanitized[:base_len] + suffix
        candidate = candidate.strip("'")
        if candidate not in existing_set:
            return candidate
    # fallback
    return sanitized[:31]


def serialize_value(field, obj):
    """Serialize a field value to an Excel-native primitive."""
    try:
        value = getattr(obj, field.name, None)
        # For FileField the attribute is a FieldFile
        if value is None:
            return None
    except Exception:
        return None

    # FileField / ImageField
    if isinstance(field, models.FileField):
        # FieldFile
        try:
            if hasattr(value, "name") and value.name:
                return value.name
            if hasattr(value, "url") and value.url:
                return value.url
        except Exception:
            pass
        return str(value) if value else None

    if isinstance(field, models.JSONField):
        if value is None:
            return None
        try:
            return json.dumps(value) if not isinstance(value, str) else value
        except Exception:
            return str(value)

    if isinstance(field, models.BooleanField):
        return bool(value)

    if isinstance(field, models.DecimalField):
        try:
            return float(value)
        except Exception:
            return str(value)

    if isinstance(field, (models.AutoField, models.BigAutoField, models.SmallAutoField)):
        try:
            return int(value)
        except Exception:
            return value

    if isinstance(field, (models.IntegerField, models.BigIntegerField, models.PositiveIntegerField, models.PositiveSmallIntegerField, models.SmallIntegerField)):
        try:
            return int(value)
        except Exception:
            return value

    if isinstance(field, models.DateTimeField):
        if isinstance(value, datetime) and value.tzinfo is not None:
            value = value.replace(tzinfo=None)
        return value

    if isinstance(field, models.DateField):
        return value

    if isinstance(field, models.TimeField):
        if isinstance(value, time) and getattr(value, "tzinfo", None) is not None:
            value = value.replace(tzinfo=None)
        return value

    # Char/Text/Slug/URL/Email and others
    return str(value) if value is not None else None


def columns_for_model(model):
    """
    Return list of (header, getter, field) for concrete fields.

    FK/OneToOne → two columns: <name>__str__ and <name>_id
    Others → one column with verbose_name.
    """
    cols = []
    for field in model._meta.concrete_fields:
        if isinstance(field, (models.ForeignKey, models.OneToOneField)) and not field.auto_created:
            # __str__ column
            def make_str_getter(f):
                def getter(obj, _f=f):
                    rel = getattr(obj, _f.name, None)
                    if rel is None:
                        return None
                    try:
                        return str(rel)
                    except Exception:
                        return str(rel)
                return getter

            def make_id_getter(f):
                def getter(obj, _f=f):
                    return getattr(obj, _f.attname, None)
                return getter

            cols.append((f"{field.name}__str__", make_str_getter(field), field))
            # header for id is attname (e.g. project_id)
            cols.append((field.attname, make_id_getter(field), field))
        else:
            header = str(getattr(field, "verbose_name", field.name)) or field.name

            def make_getter(f):
                def getter(obj, _f=f):
                    return serialize_value(_f, obj)
                return getter

            cols.append((header, make_getter(field), field))
    return cols


def get_related_targets(model):
    """Collect distinct forward FK/OneToOne target models (single hop)."""
    seen = {}
    for field in model._meta.concrete_fields:
        if isinstance(field, (models.ForeignKey, models.OneToOneField)) and not field.auto_created:
            target = field.related_model
            if target is None:
                continue
            label = target._meta.label
            if label not in seen:
                seen[label] = target
    return list(seen.values())


def _sheet_title_for_model(model):
    # Use verbose_name as sheet name base
    name = getattr(model._meta, "verbose_name", None)
    if name:
        return str(name)
    name = getattr(model._meta, "verbose_name_plural", None)
    if name:
        return str(name)
    return model._meta.model_name


def _get_or_create_sheet(wb, title, existing_names):
    sanitized = sanitize_sheet_name(title, existing_names)
    # Reuse default empty sheet if wb is fresh
    if len(wb.sheetnames) == 1 and wb.active.title == "Sheet":
        ws = wb.active
        # Check if empty (no data)
        if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None and not existing_names:
            ws.title = sanitized
            existing_names.add(sanitized)
            return ws
    ws = wb.create_sheet(title=sanitized)
    existing_names.add(sanitized)
    return ws


def style_sheet(ws):
    """Apply header styling, banded rows, freeze."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    primary = _primary_color().lstrip("#")
    if len(primary) == 6:
        argb = f"FF{primary}"
    else:
        argb = "FFC92FFF"
    try:
        # Validate hex
        int(primary, 16)
    except Exception:
        argb = "FFC92FFF"

    header_fill = PatternFill(start_color=argb, end_color=argb, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(bottom=thin)

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    # Banded rows
    band1 = PatternFill(start_color="FFF9FAFB", end_color="FFF9FAFB", fill_type="solid")
    band2 = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
    # Determine which columns are Decimal for number_format
    # We don't have column->field mapping here, but we can set number_format for numeric cells
    for r_idx in range(2, ws.max_row + 1):
        fill = band1 if r_idx % 2 == 0 else band2
        for cell in ws[r_idx]:
            cell.fill = fill
            # Apply number_format for floats that came from Decimal - set generically for floats
            if isinstance(cell.value, float):
                cell.number_format = '#,##0.00'

    ws.freeze_panes = "A2"


def autosize_columns(ws):
    """Auto-size columns capped at 50, scaled 1.2."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = cell.value
                if val is None:
                    text = ""
                else:
                    text = str(val)
                length = len(text)
                if length > max_len:
                    max_len = length
            except Exception:
                continue
        adjusted = max_len * 1.2 + 2
        if adjusted < 10:
            adjusted = 10
        if adjusted > 50:
            adjusted = 50
        ws.column_dimensions[col_letter].width = adjusted


def _write_sheet_for_model(wb, model, queryset, existing_names):
    """Write one model sheet into wb and style it."""
    columns = columns_for_model(model)
    # FK names for select_related
    fk_names = [f.name for f in model._meta.concrete_fields if isinstance(f, (models.ForeignKey, models.OneToOneField)) and not f.auto_created]

    title = _sheet_title_for_model(model)
    ws = _get_or_create_sheet(wb, title, existing_names)

    # Header - handle empty placeholder sheet (openpyxl creates 1 empty row)
    headers = [h for h, _, _ in columns]
    is_empty = ws.max_row == 1 and all(cell.value is None for cell in ws[1])
    if is_empty:
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=h)
    else:
        ws.append(headers)

    # Prepare queryset with select_related if possible
    qs = queryset
    if fk_names and hasattr(qs, "select_related"):
        try:
            qs = qs.select_related(*fk_names)
        except Exception:
            pass

    # Iterate
    iterator = qs.iterator() if hasattr(qs, "iterator") else iter(qs)
    for obj in iterator:
        row = []
        for _, getter, _ in columns:
            try:
                val = getter(obj)
            except Exception:
                val = None
            # Normalize empty string to None for blank cell? Keep as None for blank
            if val == "":
                val = None
            row.append(val)
        ws.append(row)

    style_sheet(ws)
    autosize_columns(ws)
    return wb


def build_workbook_for_queryset(model, queryset, include_related=False, existing_workbook=None):
    """Build workbook for a queryset, optionally with related sheets."""
    from openpyxl import Workbook

    wb = existing_workbook if existing_workbook is not None else Workbook()
    existing_names = set(wb.sheetnames) if existing_workbook is not None and not (len(wb.sheetnames) == 1 and wb.active.title == "Sheet" and wb.active.max_row == 1 and wb.active.cell(1, 1).value is None) else set()
    # Correct handling: if wb is fresh and default sheet empty, treat as no existing names
    if existing_workbook is None:
        # fresh workbook: treat as empty for naming
        existing_names = set()
        # But if workbook already has sheets from prior calls, existing_names is set(wb.sheetnames)
        # For existing_workbook case, use actual names except the placeholder "Sheet" if empty
        if len(wb.sheetnames) == 1 and wb.active.title == "Sheet" and wb.active.cell(1, 1).value is None:
            existing_names = set()

    wb = _write_sheet_for_model(wb, model, queryset, existing_names)

    if include_related:
        # Collect distinct targets and their FK fields
        fk_fields = [f for f in model._meta.concrete_fields if isinstance(f, (models.ForeignKey, models.OneToOneField)) and not f.auto_created]
        # Group by target
        target_to_fields = {}
        for f in fk_fields:
            target = f.related_model
            if target is None:
                continue
            target_to_fields.setdefault(target._meta.label, (target, []))[1].append(f)

        for label, (target, fields) in target_to_fields.items():
            # Collect distinct ids across all FK fields pointing to this target
            distinct_ids = set()
            for f in fields:
                try:
                    # Need fresh queryset for values_list (can't use iterator)
                    ids = queryset.values_list(f.attname, flat=True)
                    # Filter None
                    for _id in ids:
                        if _id is not None:
                            distinct_ids.add(_id)
                except Exception:
                    # Fallback: iterate objects
                    try:
                        for obj in queryset.iterator() if hasattr(queryset, "iterator") else queryset:
                            _id = getattr(obj, f.attname, None)
                            if _id is not None:
                                distinct_ids.add(_id)
                    except Exception:
                        continue
            # Fetch related queryset
            if distinct_ids:
                related_qs = target.objects.filter(pk__in=distinct_ids)
            else:
                # No referenced rows — still create sheet with header only
                related_qs = target.objects.none()
            wb = _write_sheet_for_model(wb, target, related_qs, existing_names)

    # Remove placeholder Sheet if still present and we have other sheets
    if len(wb.sheetnames) > 1 and "Sheet" in wb.sheetnames:
        try:
            ws = wb["Sheet"]
            if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
                wb.remove(ws)
        except Exception:
            pass

    return wb


def build_full_app_workbook():
    """Build workbook with one sheet per ourlives model, all rows."""
    from openpyxl import Workbook

    wb = Workbook()
    existing_names = set()
    # Remove initial handling will be done by _write_sheet

    ourlives_models = [m for m in apps.get_models() if m._meta.app_label == "ourlives"]
    ourlives_models = sorted(ourlives_models, key=lambda m: m._meta.model_name)

    for model in ourlives_models:
        try:
            qs = model.objects.all()
        except Exception:
            continue
        wb = _write_sheet_for_model(wb, model, qs, existing_names)

    if len(wb.sheetnames) > 1 and "Sheet" in wb.sheetnames:
        try:
            ws = wb["Sheet"]
            if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
                wb.remove(ws)
        except Exception:
            pass

    return wb
